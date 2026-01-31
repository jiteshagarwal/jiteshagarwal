import random
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

# --- CONFIGURATION ---
filename = "Smart_Availability_Matrix.xlsx"
regions = ["USA", "Europe", "UAE", "APAC", "Saudi"]
languages = ["English", "Hindi", "Spanish", "German", "Arabic", "French", "Italian"]
services = ["ASR", "NLP Services", "Gen AI Services", "KaaS/PKaaS"]
products = ["SSA", "RTGA", "CIA", "CRA"]
statuses = ["Full Support", "Limited Support", "Not Supported"]

# Dependency Logic (For Reference/Comments)
# SSA: ASR, Gen AI, KaaS
# RTGA: ASR, NLP, Gen AI, KaaS
# CIA: NLP, Gen AI
# CRA: ASR

wb = Workbook()

# ==========================================
# SHEET 1: SERVICE INPUT (The Source of Truth)
# ==========================================
ws_data = wb.active
ws_data.title = "Service_Input"
headers = ["Region", "Service", "Language", "Status", "Lookup_Key"]
ws_data.append(headers)

# Formatting
header_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for cell in ws_data[1]:
    cell.fill = header_fill
    cell.font = header_font

# Generate Data (Full Factorial for SERVICES only)
row_num = 2
for r in regions:
    for s in services:
        for l in languages:
            # Random Status
            status = random.choices(statuses, weights=[40, 30, 30], k=1)[0]
            # Key: Region|Service|Language
            key_formula = f'=A{row_num}&"|"&B{row_num}&"|"&C{row_num}'
            ws_data.append([r, s, l, status, key_formula])
            row_num += 1

ws_data.column_dimensions['E'].hidden = True

# ==========================================
# STYLING HELPERS
# ==========================================
def apply_dark_theme_formatting(ws, start_row, start_col, num_cols, num_rows):
    # Background
    ws.sheet_view.showGridLines = False
    
    # Conditional Formatting Ranges
    matrix_range = f"{get_column_letter(start_col)}{start_row+1}:{get_column_letter(start_col+num_cols-1)}{start_row+num_rows}"

    # Green (Full Support)
    green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    dxf_green = DifferentialStyle(fill=green_fill, font=Font(color="FFFFFF"))
    rule_green = Rule(type="containsText", operator="containsText", text="Full Support", dxf=dxf_green)
    rule_green.formula = [f'NOT(ISERROR(SEARCH("Full Support", {matrix_range})))']
    ws.conditional_formatting.add(matrix_range, rule_green)

    # Yellow (Limited)
    yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    dxf_yellow = DifferentialStyle(fill=yellow_fill) 
    rule_yellow = Rule(type="containsText", operator="containsText", text="Limited Support", dxf=dxf_yellow)
    rule_yellow.formula = [f'NOT(ISERROR(SEARCH("Limited Support", {matrix_range})))']
    ws.conditional_formatting.add(matrix_range, rule_yellow)

    # Red (Not Supported)
    red_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    dxf_red = DifferentialStyle(fill=red_fill, font=Font(color="FFFFFF"))
    rule_red = Rule(type="containsText", operator="containsText", text="Not Supported", dxf=dxf_red)
    rule_red.formula = [f'NOT(ISERROR(SEARCH("Not Supported", {matrix_range})))']
    ws.conditional_formatting.add(matrix_range, rule_red)

# ==========================================
# SHEET 2: SERVICE DASHBOARD
# ==========================================
ws_serv = wb.create_sheet("Service_Dashboard")

# Header Controls
ws_serv['B2'] = "SERVICE AVAILABILITY MATRIX"
ws_serv['B2'].font = Font(size=16, bold=True, color="203764")
ws_serv['B4'] = "Select Region:"
ws_serv['C4'] = regions[0] # Default
ws_serv['C4'].font = Font(bold=True)
ws_serv['C4'].border = Border(bottom=Side(style='thin'))

# Dropdown
dv = DataValidation(type="list", formula1=f'"{",".join(regions)}"', allow_blank=False)
ws_serv.add_data_validation(dv)
dv.add(ws_serv['C4'])

# Grid Setup
start_row = 7
start_col = 2
ws_serv.cell(row=start_row, column=1, value="Language").font = Font(bold=True)

# Headers (Services)
for i, s in enumerate(services):
    c = start_col + i
    cell = ws_serv.cell(row=start_row, column=c, value=s)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="2F5597", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')
    ws_serv.column_dimensions[get_column_letter(c)].width = 20

# Rows (Languages) & Logic
for i, lang in enumerate(languages):
    r = start_row + 1 + i
    ws_serv.cell(row=r, column=1, value=lang).font = Font(bold=True)
    
    for j, s in enumerate(services):
        c = start_col + j
        # INDEX MATCH to look up Service Status directly
        formula = f'=INDEX(Service_Input!$D:$D, MATCH($C$4&"|"&{get_column_letter(c)}${start_row}&"|"&$A{r}, Service_Input!$E:$E, 0))'
        cell = ws_serv.cell(row=r, column=c, value=formula)
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

apply_dark_theme_formatting(ws_serv, start_row, start_col, len(services), len(languages))

# ==========================================
# SHEET 3: PRODUCT DASHBOARD (The Smart One)
# ==========================================
ws_prod = wb.create_sheet("Product_Dashboard")

# Header Controls
ws_prod['B2'] = "PRODUCT AVAILABILITY MATRIX (Auto-Calculated)"
ws_prod['B2'].font = Font(size=16, bold=True, color="203764")
ws_prod['B4'] = "Select Region:"
ws_prod['C4'] = regions[0] 
ws_prod['C4'].font = Font(bold=True)
ws_prod['C4'].border = Border(bottom=Side(style='thin'))

# Dropdown
ws_prod.add_data_validation(dv) # Use same validation
dv.add(ws_prod['C4'])

# Legend Note
ws_prod['E4'] = "*Calculated based on dependencies (e.g. SSA requires ASR, GenAI, KaaS)"
ws_prod['E4'].font = Font(italic=True, size=9, color="555555")

# Grid Setup
start_row = 7
start_col = 2
ws_prod.cell(row=start_row, column=1, value="Language").font = Font(bold=True)

# Headers (Products)
for i, p in enumerate(products):
    c = start_col + i
    cell = ws_prod.cell(row=start_row, column=c, value=p)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="203764", fill_type="solid") # Darker blue for products
    cell.alignment = Alignment(horizontal='center')
    ws_prod.column_dimensions[get_column_letter(c)].width = 20

# HELPER FUNCTION FOR FORMULA GENERATION
def get_lookup_formula(service_name, row_idx):
    # Returns the INDEX/MATCH part for a specific service
    return f'INDEX(Service_Input!$D:$D, MATCH($C$4&"|{service_name}|"&$A{row_idx}, Service_Input!$E:$E, 0))'

# Rows (Languages) & DEPENDENCY LOGIC
for i, lang in enumerate(languages):
    r = start_row + 1 + i
    ws_prod.cell(row=r, column=1, value=lang).font = Font(bold=True)
    
    # 1. SSA (ASR, Gen AI, KaaS)
    # Logic: If ANY are "Not Supported" -> Not Supported. Else if ANY are "Limited" -> Limited. Else Full.
    f_asr = get_lookup_formula("ASR", r)
    f_gen = get_lookup_formula("Gen AI Services", r)
    f_kaas = get_lookup_formula("KaaS/PKaaS", r)
    
    formula_ssa = (
        f'=IF(OR({f_asr}="Not Supported", {f_gen}="Not Supported", {f_kaas}="Not Supported"), "Not Supported", '
        f'IF(OR({f_asr}="Limited Support", {f_gen}="Limited Support", {f_kaas}="Limited Support"), "Limited Support", "Full Support"))'
    )
    ws_prod.cell(row=r, column=2, value=formula_ssa).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 2. RTGA (ASR, NLP, Gen AI, KaaS)
    f_nlp = get_lookup_formula("NLP Services", r)
    formula_rtga = (
        f'=IF(OR({f_asr}="Not Supported", {f_nlp}="Not Supported", {f_gen}="Not Supported", {f_kaas}="Not Supported"), "Not Supported", '
        f'IF(OR({f_asr}="Limited Support", {f_nlp}="Limited Support", {f_gen}="Limited Support", {f_kaas}="Limited Support"), "Limited Support", "Full Support"))'
    )
    ws_prod.cell(row=r, column=3, value=formula_rtga).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 3. CIA (NLP, Gen AI)
    formula_cia = (
        f'=IF(OR({f_nlp}="Not Supported", {f_gen}="Not Supported"), "Not Supported", '
        f'IF(OR({f_nlp}="Limited Support", {f_gen}="Limited Support"), "Limited Support", "Full Support"))'
    )
    ws_prod.cell(row=r, column=4, value=formula_cia).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 4. CRA (ASR)
    formula_cra = (
        f'=IF({f_asr}="Not Supported", "Not Supported", '
        f'IF({f_asr}="Limited Support", "Limited Support", "Full Support"))'
    )
    ws_prod.cell(row=r, column=5, value=formula_cra).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

apply_dark_theme_formatting(ws_prod, start_row, start_col, len(products), len(languages))
ws_prod.column_dimensions['A'].width = 15
ws_serv.column_dimensions['A'].width = 15

# Save
wb.save(filename)
print(f"File '{filename}' created successfully.")